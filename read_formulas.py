import openpyxl
import sys

file_path = 'Codifica (Productos Varios prueba).xlsm'
try:
    wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=False)
except Exception as e:
    print('Error load:', e)
    sys.exit(1)

def explore_formulas(sheet_name):
    print(f'\n--- FORMULAS {sheet_name} ---')
    if sheet_name not in wb.sheetnames:
        print('no existe')
        return
    ws = wb[sheet_name]
    row_count = 0
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=15):
        formulas = []
        for cell in row:
            if cell.data_type == 'f': # Formula
                formulas.append(f'{cell.coordinate}: {cell.value}')
        if formulas:
            print(' | '.join(formulas))

for s in ['PARA CONSULTA', 'TABLA RECETA', 'merma y dolar proyec', 'Precio por Formato']:
    explore_formulas(s)
